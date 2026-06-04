#!/usr/bin/env python3
"""Convert Networking groups xlsx → email,name CSV for import scripts."""
import re
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Run: pip3 install openpyxl")
    sys.exit(1)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", re.I)

SHEET_CONFIG = {
    "Networking groups plus emails 2": {"name": 0, "email": 3, "phone": 4, "skip_header": True},
    "New": {"name": 0, "email": 1, "skip_header": True},
    "MasterClasses": {"name": 0, "email": 1, "skip_header": False},
    "Exhibitions": {"name": 0, "email": 1, "skip_header": True},
}


def extract_emails(cell):
    if cell is None:
        return []
    text = str(cell).strip()
    if not text or text.lower() in ("none", "email", "contact email", "email address"):
        return []
    found = EMAIL_RE.findall(text)
    seen = set()
    out = []
    for e in found:
        el = e.lower()
        if el not in seen:
            seen.add(el)
            out.append(el)
    return out


def clean_name(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("none", "organiser name", "networking group / organisation", "exhibition / event"):
        return ""
    return s


def rows_from_sheet(ws, cfg):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if cfg.get("skip_header") and i == 0:
            continue
        name = clean_name(row[cfg["name"]] if len(row) > cfg["name"] else None)
        email_cell = row[cfg["email"]] if len(row) > cfg["email"] else None
        phone = None
        if "phone" in cfg and len(row) > cfg["phone"]:
            p = row[cfg["phone"]]
            if p is not None and str(p).strip().lower() != "none":
                phone = str(p).strip()
        for email in extract_emails(email_cell):
            if not name:
                name = email.split("@")[0].replace(".", " ").title()
            rows.append({"email": email, "name": name, "phone": phone or ""})
    return rows


def directory_ops_sheet(ws):
    rows = []
    header_idx = None
    all_rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(all_rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if "business name" in cells and "email address" in cells:
            header_idx = i
            name_i = cells.index("business name")
            email_i = cells.index("email address")
            phone_i = cells.index("phone number") if "phone number" in cells else None
            for row2 in all_rows[i + 1 :]:
                if not row2 or not any(row2):
                    continue
                name = clean_name(row2[name_i] if len(row2) > name_i else None)
                for email in extract_emails(row2[email_i] if len(row2) > email_i else None):
                    phone = ""
                    if phone_i is not None and len(row2) > phone_i and row2[phone_i]:
                        phone = str(row2[phone_i]).strip()
                    if not name:
                        name = email.split("@")[0].replace(".", " ").title()
                    rows.append({"email": email, "name": name, "phone": phone})
            break
    return rows


def main():
    xlsx = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else "/Users/catherinehancher/Desktop/Networking groups-3:6.xlsx"
    )
    out_dir = Path(__file__).resolve().parent.parent / "data"
    out_dir.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    combined = []
    seen_email = set()

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue
        for r in rows_from_sheet(wb[sheet_name], cfg):
            if r["email"] not in seen_email:
                seen_email.add(r["email"])
                combined.append(r)

    if "Directory Ops" in wb.sheetnames:
        for r in directory_ops_sheet(wb["Directory Ops"]):
            if r["email"] not in seen_email:
                seen_email.add(r["email"])
                combined.append(r)

    wb.close()

    attendees_csv = out_dir / "networking-groups-import.csv"
    organisers_csv = out_dir / "networking-groups-organisers.csv"

    with attendees_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["email", "name"])
        w.writeheader()
        for r in combined:
            w.writerow({"email": r["email"], "name": r["name"]})

    with organisers_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["email", "name", "phone"])
        w.writeheader()
        for r in combined:
            w.writerow({"email": r["email"], "name": r["name"], "phone": r.get("phone", "")})

    print(f"Wrote {len(combined)} rows to {attendees_csv}")
    print(f"Wrote {len(combined)} rows to {organisers_csv}")


if __name__ == "__main__":
    main()
